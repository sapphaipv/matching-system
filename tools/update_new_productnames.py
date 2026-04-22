from openpyxl import load_workbook
import pandas as pd
import os

# 1. Cấu hình đường dẫn

store_name = "205NTT"
input_dir = r"D:\CTY\TenSanPham\\" + store_name

file_sp = os.path.join(input_dir, "SanPham.xlsx")
file_new = os.path.join(input_dir, "TenHang_New.xlsx")
output_file = os.path.join(input_dir, "SanPham_Updated.xlsx")
log_file = os.path.join(input_dir, "KQ_Updated.xlsx")

def update_keep_100_format():

    col_ma = "Mã hàng"
    col_ten = "Tên hàng"
    col_nhom = "Nhóm hàng(3 Cấp)"

    PREFIX = "_SỬA_TÊN_"

    # ===== 1. Đọc mapping =====
    df_new = pd.read_excel(file_new)
    df_new[col_ma] = df_new[col_ma].astype(str).str.strip()
    df_new[col_ten] = df_new[col_ten].astype(str).str.strip()

    mapping = dict(zip(df_new[col_ma], df_new[col_ten]))

    # ===== 2. Mở Excel gốc =====
    wb = load_workbook(file_sp)
    ws = wb.active

    # ===== 3. Xác định cột =====
    headers = [cell.value for cell in ws[1]]

    for col in [col_ma, col_ten, col_nhom]:
        if col not in headers:
            print(f"❌ Không tìm thấy cột: {col}")
            return

    col_ma_idx = headers.index(col_ma) + 1
    col_ten_idx = headers.index(col_ten) + 1
    col_nhom_idx = headers.index(col_nhom) + 1

    # ===== 4. Update + log =====
    updated = 0
    sua_ten_count = 0

    log_data = []  # lưu log

    for row in ws.iter_rows(min_row=2):
        cell_ma = row[col_ma_idx - 1]
        cell_ten = row[col_ten_idx - 1]
        cell_nhom = row[col_nhom_idx - 1]

        if cell_ma.value is None:
            continue

        ma = str(cell_ma.value).strip()

        if ma in mapping:
            new_name = mapping[ma]
            old_name = str(cell_ten.value).strip() if cell_ten.value else ""

            if old_name != new_name:
                # ===== ghi log =====
                log_data.append({
                    "Tên hàng (mới)": new_name,
                    "Mã hàng": ma,
                    "Tên hàng cũ": old_name
                })

                # ===== update =====
                cell_ten.value = new_name
                updated += 1

                # ===== xử lý prefix =====
                if cell_nhom.value:
                    nhom_val = str(cell_nhom.value).strip()

                    if nhom_val.startswith(PREFIX):
                        cell_nhom.value = nhom_val.replace(PREFIX, "", 1)
                        sua_ten_count += 1

    # ===== 5. Lưu file Excel chính =====
    wb.save(output_file)

    # ===== 6. Xuất log =====
    if log_data:
        df_log = pd.DataFrame(log_data)
        df_log.to_excel(log_file, index=False)
    else:
        print("⚠️ Không có dòng nào được update")

    # ===== 7. Thông báo =====
    print("✔ Hoàn tất")
    print(f"✔ Tổng dòng được update: {updated}")
    print(f"✔ Trong đó có {sua_ten_count} dòng có '{PREFIX}'")
    print(f"✔ File log: {log_file}")


if __name__ == "__main__":
    update_keep_100_format()
